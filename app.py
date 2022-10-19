from io import BytesIO
from flask import Flask, render_template, url_for, request, redirect, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from sqlalchemy.orm import relationship
from sqlalchemy import Column, Integer, String, Text, DateTime, ForeignKey
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import openpyxl as exl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from wtforms.validators import ValidationError
from libs import cost_center_num, types, names


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
engine = create_engine('sqlite:///database.db', echo=True)
Session = sessionmaker(bind=engine)
session = Session()



class Article(db.Model):
    __tablename__ = 'article'
    id = Column(Integer, primary_key=True)
    requis = Column(String(50))
    vendor = Column(Text, nullable=False)
    desc = Column(Text, nullable=False)
    invoice = Column(Text)
    date = Column(DateTime, default=datetime.utcnow)
    piar = relationship('Piar', backref='article', lazy='dynamic')


class Piar(db.Model):
    __tablename__ = 'piar'
    id = Column(Integer, primary_key=True)
    unit_price = Column(Integer, nullable=False)
    unit_desc = Column(Text, nullable=False)
    remark = Column(Text, nullable=False)
    quality = Column(Integer, nullable=False)
    respon = Column(String(100), nullable=False)
    cost = Column(Text, nullable=False)
    type = Column(Text, nullable=False)
    article_id = Column(Integer, ForeignKey('article.id'))


@app.route('/index')
@app.route('/')
def index():
    articles = Article.query.order_by(Article.id.desc()).all()
    return render_template('index.html', articles=articles)


@app.route('/create', methods=['POST', 'GET'])
def create():
    if request.method == 'POST':
        id = request.form.get('id')
        desc = request.form['desc']
        vendor = request.form['vendor']
        date = request.form.get('date')
        requis = request.form['requis']
        invoice = request.form['invoice']

        article = Article(id=id,
                          invoice=invoice,
                          requis=requis,
                          desc=desc,
                          vendor=vendor,
                          date=date)

        session.add(article)
        session.commit()
        return redirect('/index')

    else:
        articles = Article.query.all()

        return render_template('create.html',  names=names, articles=articles)

@app.route('/support', methods = ['POST', 'GET'])
def support():
    if request.method == 'POST':
        add = request.form['support_add_name']
        del_name = request.form['support_del_name']
        try:
            names.remove(f'{del_name}')
            names.append(add)
        except:
            pass
    else:
        pass
    return render_template('support.html', names=names)
@app.route('/index/<int:id>/pr', methods=['POST', 'GET'])
def getPR(id):
    if request.method == "POST":
        unit_price = request.form['unit_price']
        unit_desc = request.form['unit_desc']
        remark = request.form['remark']
        quality = request.form['quality']
        respon = request.form['respon']
        cost = request.form['cost']
        type = request.form['type']

        piar = Piar(
            unit_price=unit_price,
            unit_desc=unit_desc,
            remark=remark,
            quality=quality,
            respon=respon,
            cost=cost,
            article_id=id,
            type=type)


        session.add(piar)
        session.commit()
        return '<script>document.location.href = document.referrer</script>'

    else:
        piars = Piar.query.filter(Piar.article_id.in_([id])).all()

        return render_template('pr.html', piars=piars, cost_center_num=cost_center_num.keys(), types=types.keys())

@app.route('/index/1/pr/<int:id>/del')
def pr_delete(id):
    piar = Piar.query.get_or_404(id)

    db.session.delete(piar)
    db.session.commit()

    return '<script>document.location.href = document.referrer</script>'
# config openpyxl
wb = exl.load_workbook(filename="test.xlsx")
ws = wb.active
sheet = wb["BSG-PR"]
font_style = Font(name='Verdana', sz='11')
alig_style_left = Alignment(wrapText=True, horizontal='left', vertical='center')
alig_style_right = Alignment(wrapText=True, horizontal='right', vertical='center')
alig_style_center = Alignment(wrapText=True, horizontal='center', vertical='center')

border = Border(top=Side(border_style='thin'),
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                bottom=Side(border_style='thin'))

border_for_test_left = Border(top=Side(border_style='thin', color='A6A6A6'),
                             left=Side(border_style='thin', color='A6A6A6'),
                              bottom=Side(border_style='thin', color='A6A6A6'))

border_for_test_right = Border(top=Side(border_style='thin', color='A6A6A6'),
                                 right=Side(border_style='thin', color='A6A6A6'),
                                  bottom=Side(border_style='thin', color='A6A6A6'))

border_bottom_top_test = Border(bottom=Side(border_style='thin', color='A6A6A6'),
                                top=Side(border_style='thin', color='A6A6A6'))

border_bottom_top = Border(bottom=Side(border_style='thin'),
                           top=Side(border_style='thin'))


@app.route('/index/<int:id>', methods=['GET', 'POST'])
def getCSV(id):
        id = [id]
        value = Article.query.filter(Article.id.in_(id)).all()
        piar_value = Piar.query.filter(Piar.article_id.in_(id)).all()
        last_value = len(piar_value)
        ws.merge_cells(f'I18:I{17 + last_value}')
        ws[f'I{17 + last_value}'].border = Border(bottom=Side(border_style='thin'))

        for i, lx in enumerate(piar_value):
            ws.merge_cells(f'C{18 + i}:D{18 + i}')
            ws.merge_cells(f'F{18 + i}:G{18 + i}')
            ws.merge_cells(f'J{18 + i}:K{18 + i}')

            ws[f'N{18 + i}'].border = border
            ws[f'N{18 + i}'].value = f'=SUM(E{18 + i}*M{18 + i})'
            ws[f'N{18 + i}'].number_format = '#,##0.00 KZT'
            ws[f'N{18 + i}'].font = font_style
            ws.row_dimensions[18 + i].height = 71
            ws[f'N{18 + i}'].alignment = alig_style_center

            ws[f'D{18 + i}'].border = border_bottom_top
            ws[f'G{18 + i}'].border = border_bottom_top
            ws[f'K{18 + i}'].border = border_bottom_top

            ws[f'B{18 + i}'].border = border
            ws[f'B{18 + i}'].value = 1+i
            ws[f'B{18 + i}'].font = font_style
            ws.row_dimensions[18 + i].height = 71
            ws[f'B{18 + i}'].alignment = alig_style_center

            ws[f'C{18+i}'].border = border
            ws[f'C{18+i}'].value = lx.unit_desc
            ws[f'C{18 + i}'].font = font_style
            ws.row_dimensions[18+i].height = 71
            ws[f'C{18 + i}'].alignment = alig_style_center

            ws[f'E{18 + i}'].border = border
            ws[f'E{18 + i}'].value = lx.quality
            ws[f'E{18 + i}'].font = font_style
            ws.row_dimensions[18 + i].height = 71
            ws[f'E{18 + i}'].alignment = alig_style_center

            ws[f'F{18 + i}'].border = border
            ws[f'F{18 + i}'].value = lx.respon
            ws[f'F{18 + i}'].font = font_style
            ws.row_dimensions[18 + i].height = 71
            ws[f'F{18 + i}'].alignment = alig_style_center

            ws[f'L{18 + i}'].border = border
            ws[f'L{18 + i}'].value = f'{cost_center_num[lx.cost]}-{types[lx.type]}'
            ws[f'L{18 + i}'].font = font_style
            ws.row_dimensions[18 + i].height = 71
            ws[f'L{18 + i}'].alignment = alig_style_center

            ws[f'M{18 + i}'].border = border
            ws[f'M{18 + i}'].number_format = '#,##0.00 KZT'
            ws[f'M{18 + i}'].value = lx.unit_price
            ws[f'M{18 + i}'].font = font_style
            ws.row_dimensions[18 + i].height = 71
            ws[f'M{18 + i}'].alignment = alig_style_center

            ws[f'H{18 + i}'].border = border
            ws[f'H{18 + i}'].value = 'ea'
            ws[f'H{18 + i}'].font = font_style
            ws.row_dimensions[18 + i].height = 71
            ws[f'H{18 + i}'].alignment = alig_style_center

            ws[f'J{18 + i}'].border = border
            ws[f'J{18 + i}'].value = 'ea'
            ws[f'J{18 + i}'].font = font_style
            ws.row_dimensions[18 + i].height = 71
            ws[f'J{18 + i}'].alignment = alig_style_center

        for xl in value:
            ws['E4'] = 'SVR-PR-' + '{:05}'.format(xl.id)
            ws['I18'].value = xl.vendor
            ws['I4'].border = border_bottom_top_test
            ws['I4'].value = xl.requis
            ws['E5'].value = xl.date

            ws['I18'].border = border_bottom_top_test
            ws['I2'].border = border_bottom_top_test
            ws['J2'].border = border_bottom_top_test
            ws['K2'].border = border_bottom_top_test
            ws['L2'].border = border_bottom_top_test
            ws['M2'].border = border_bottom_top_test
            ws['N2'].border = border_bottom_top_test
            ws['N2'].border = border_bottom_top_test
            ws['O2'].border = border_for_test_right
            ws['I4'].border = border_bottom_top_test
            ws['J4'].border = border_bottom_top_test
            ws['K4'].border = border_bottom_top_test
            ws['L4'].border = border_bottom_top_test
            ws['M4'].border = border_bottom_top_test
            ws['N4'].border = border_bottom_top_test
            ws['N4'].border = border_for_test_right
            save_name = 'SVR-PR-' + '{:05}'.format(xl.id)

        ws.merge_cells(f'B{18 + last_value}:M{18 + last_value}')
        ws[f'B{18 + last_value}'].value = 'Total'
        ws[f'B{18 + last_value}'].alignment = alig_style_center
        ws.row_dimensions[18 + last_value].height = 13
        ws[f'B{18 + last_value}'].font = Font(bold=True, name='Verdana', sz='11')
        ws[f'B{18 + last_value}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        ws[f'N{18 + last_value}'].value = f'=SUM(N18:N{17 + last_value})'
        ws[f'N{18 + last_value}'].alignment = alig_style_center
        ws[f'N{18 + last_value}'].font = font_style
        ws[f'N{18 + last_value}'].border = border
        ws[f'N{18 + last_value}'].number_format =  '#,##0.00 KZT'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, download_name=f"{save_name}.xlsx", as_attachment=True)

if __name__=='__main__':
    app.run(host='0.0.0.0', debug=True)