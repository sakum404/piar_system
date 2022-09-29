import openpyxl as exl

wb = exl.Workbook()
ws = wb.active
sheet = wb["Sheet"]

# пустой список
data = []


# даем значение к списку
data1 = 1
data2 = 2
data3 = None


# добавляем все значение к списку
data.append(data1)
data.append(data2)
data.append(data3)

# фильтруем пустые значние
out = filter(None, data)

# добавляем данные из списка в excel
for i, statdata1 in enumerate(data):
    ws.cell(row=1+i, column=1).value = statdata1

# сохраняем excel файл
wb.save('tese.xlsx')

# находим последнюю пустую ячейку
a1 = sheet.max_row

# добавляем к пустой ячейке суммированный значение из списка
ws[f'A{a1}'].value = sum(out)












